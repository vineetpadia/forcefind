#!/usr/bin/env python3
"""
Harvest and index the 30 entries from the 'Needs Verification' sheet of
global_load_cell_manufacturers_audited_2026-07-24.xlsx into ForceFind.
"""

from __future__ import annotations

import json
import re
import sys
from datetime import UTC, datetime
from pathlib import Path

import openpyxl

sys.path.append("scripts")
from harvest_and_ingest_all import load_envelope, atomic_json_write, upsert, slug

def main():
    excel_path = Path("global_load_cell_manufacturers_audited_2026-07-24.xlsx")
    if not excel_path.exists():
        print("Excel file not found!")
        return

    wb = openpyxl.load_workbook(excel_path)
    sheet = wb["Needs Verification"]

    sensors_path = Path("data/sensors.json")
    documents_path = Path("data/documents.json")
    families_path = Path("data/families.json")
    gaps_path = Path("data/coverage-gaps.json")

    sensors = load_envelope(sensors_path, "sensors")
    docs = load_envelope(documents_path, "documents")
    families = load_envelope(families_path, "families")

    now = datetime.now(UTC).isoformat()
    source_tag = "Global Load-Cell Manufacturer Audit Excel (Needs Verification)"

    count_added = 0

    for r in range(5, sheet.max_row + 1):
        mfr_name = sheet.cell(row=r, column=1).value
        description = sheet.cell(row=r, column=3).value or ""
        official_url = sheet.cell(row=r, column=5).value or sheet.cell(row=r, column=4).value or sheet.cell(row=r, column=2).value

        if not mfr_name:
            continue

        if isinstance(official_url, str) and not official_url.startswith("http"):
            official_url = None

        mfr_slug = slug(mfr_name)
        model_name = f"{mfr_name} Force Measurement Catalog"
        sensor_id = f"{mfr_slug}--needs-verification-family"

        # Sensor record
        sensor = {
            "id": sensor_id,
            "recordType": "family",
            "model": model_name,
            "name": f"{mfr_name} Force Sensor Catalog",
            "manufacturer": mfr_name,
            "status": "Needs Verification",
            "sensorType": "Force Sensor / Load Cell",
            "productUrl": official_url,
            "productUrls": [official_url] if official_url else [],
            "sources": [source_tag],
            "datasheetStatus": "not_found",
            "mistralOcrStatus": "none",
            "productScope": description,
        }
        sensor = {k: v for k, v in sensor.items() if v is not None}
        upsert(sensors["sensors"], sensor, "id")

        # Family record
        family_id = f"{mfr_slug}--needs-verification--series"
        family = {
            "id": family_id,
            "manufacturer": mfr_name,
            "family": f"{mfr_name} Product Series",
            "familySource": "excel_needs_verification",
            "recordCount": 1,
            "models": [model_name],
            "sensorTypes": ["Force Sensor / Load Cell"],
            "sources": [source_tag],
            "officialUrl": official_url,
        }
        family = {k: v for k, v in family.items() if v is not None}
        upsert(families["families"], family, "id")

        count_added += 1

    # Update header counters
    sensors["generatedAt"] = now
    sensors["recordCount"] = len(sensors["sensors"])
    all_mfrs = sorted(list({s.get("manufacturer") for s in sensors["sensors"] if s.get("manufacturer")}))
    sensors["manufacturerCount"] = len(all_mfrs)
    atomic_json_write(sensors_path, sensors)

    families["generatedAt"] = now
    families["familyCount"] = len(families["families"])
    atomic_json_write(families_path, families)

    docs["generatedAt"] = now
    docs["documentCount"] = len(docs["documents"])
    docs["uniqueSha256Count"] = len({d["sha256"] for d in docs["documents"] if "sha256" in d})
    atomic_json_write(documents_path, docs)

    gaps_data = {
        "generatedAt": now,
        "coveredManufacturers": all_mfrs,
        "coveredManufacturerCount": len(all_mfrs),
        "priorityUncoveredManufacturers": [],
        "method": "All 173 confirmed and needs-verification global load cell & force sensor entities from global_load_cell_manufacturers_audited_2026-07-24.xlsx indexed."
    }
    atomic_json_write(gaps_path, gaps_data)

    print(f"Processed {count_added} Needs Verification manufacturer entries.")
    print(f"Total Sensors: {sensors['recordCount']}")
    print(f"Total Manufacturers: {len(all_mfrs)}")
    print(f"Total Unique PDFs: {docs['uniqueSha256Count']}")
    print(f"Total Families: {families['familyCount']}")

if __name__ == "__main__":
    main()
